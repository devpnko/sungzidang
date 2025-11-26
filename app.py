import streamlit as st
import google.generativeai as genai
from supabase import create_client, Client
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import io
import time
import uuid
import pandas as pd
import random
import os
import traceback
import re

# --- Reference Data Loading ---
def load_reference_data():
    """Loads reference data (models, plans) from JSON file."""
    try:
        with open('data/reference_db.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data
    except FileNotFoundError:
        return {"models": [], "plans": []}

REFERENCE_DATA = load_reference_data()
VALID_MODEL_NAMES = [m['name'] for m in REFERENCE_DATA.get('models', [])]
VALID_PLAN_NAMES = REFERENCE_DATA.get('plans', [])

# --- 유틸리티: 랜덤 파스텔 색상 생성 (어두운 색 방지) ---
def get_random_pastel_color():
    # R, G, B를 각각 200~255 사이에서 뽑아서 무조건 밝은 색이 나오게 함
    r = lambda: random.randint(200, 255)
    return '#%02X%02X%02X' % (r(), r(), r())

# --- 데이터 구조 클래스 ---
class PolicyData:
    def __init__(self, name, image_bytes, color_hex):
        self.name = name
        self.image_bytes = image_bytes  # 원본 이미지 저장 (AI 분석은 나중에)
        self.color_hex = color_hex
        # 분석 결과는 나중에 채워짐
        self.df = None
        self.footer_text = None
        self.is_analyzed = False

# --- 1. Gemini 파싱 함수 (배틀용) ---
def parse_image_with_gemini_v2(file_bytes, agency_name, color_hex, api_key, model_name):
    """V2 전용: 배틀 모드에서 사용하는 Gemini 파싱 함수"""
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel(model_name)
    
    # Reference data 로드
    model_list_str = ", ".join(VALID_MODEL_NAMES) if VALID_MODEL_NAMES else "None"
    plan_list_str = ", ".join(VALID_PLAN_NAMES) if VALID_PLAN_NAMES else "None"
    
    prompt = f"""
    Analyze this mobile phone price sheet image FULLY from TOP to BOTTOM.
    There are often MULTIPLE tables (e.g., Premium models at top, Low-cost models at bottom).
    
    **CRITICAL Instructions:**
    1. **Scan the ENTIRE image**: Look for all tables (Main, Low Cost, etc.).
    
    2. **Header Analysis (Sub-Agency & Condition)**:
       - **Sub-Agency Detection**:
         - Look for codes like "I", "J", "K", "Eren", "Hong", etc. attached to headers (e.g., "SK-I", "KT-J").
         - If found, extract "I", "J", "Eren" as the **Sub-Agency**.
         - If NOT found (e.g., just "MNP"), use "Common" or "Main".
         
       - **Condition Detection (CRITICAL)**:
         - You MUST combine **Contract Type** + **Join Type**.
         - **Contract Type**: Look for "공시", "공시지원금" -> **"공시"**. Look for "선약", "선택약정" -> **"선약"**. (If neither found, infer from context or default to "공시").
         - **Join Type**: Look for "MNP", "번이" -> **"MNP"**. Look for "기변", "기기변경" -> **"기변"**.
         - **Output Example**: "공시 MNP", "선약 기변", "공시 신규"
         
       - **Plan Detection (CRITICAL)**:
         - Detect plan name accurately. Map to: {plan_list_str}
         - **Inference from Price**: If header has "109", "109000" -> **"5GX 프리미엄"**. If "89", "89000" -> **"5GX 프라임"**.
         - **IMPORTANT**: For "T우주", use the full name **"5GX 프리미엄(T우주)"**.
         - If no plan, use "Standard".

    3. **Footer & Conditions**:
       - Extract **ALL** text at the bottom of the image (subscription conditions, notices, additional fees, etc.).
       - Do NOT summarize. Capture the full text as a single string.

    4. **Output Format (JSON Structure)**:
       - Return a SINGLE JSON object.
       - **"columns"**: A list of objects describing each column (excluding Model column).
         - Example: `[{{"sub_agency": "I", "condition": "MNP", "plan": "5GX Prime"}}, {{"sub_agency": "J", "condition": "기변", "plan": "Save Plan"}}]`
       - **"rows"**: List of rows. Each row starts with Model Name, followed by prices corresponding to "columns".
       - **"footer"**: The extracted footer text.
       
    **Example Output:**
    {{
      "columns": [
        {{"sub_agency": "I", "condition": "MNP", "plan": "5GX Prime"}},
        {{"sub_agency": "I", "condition": "기변", "plan": "5GX Prime"}},
        {{"sub_agency": "J", "condition": "MNP", "plan": "Save Plan"}}
      ],
      "rows": [
        ["SM-S921N", 10, 20, null],
        ["SM-A245N", null, null, 0]
      ],
      "footer": "..."
    }}
    """
    
    # Safety Settings: 모든 필터 해제 (시세표가 스팸/상업적으로 분류될 수 있음)
    safety_settings = [
        {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_NONE"},
        {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_NONE"},
        {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_NONE"},
        {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_NONE"},
    ]
    
    response = model.generate_content(
        [prompt, {"mime_type": "image/jpeg", "data": file_bytes}],
        safety_settings=safety_settings
    )
    text = response.text
    print(f"DEBUG: Gemini Response Text: '{text}'") # 디버깅용 출력
    
    try:
        # 정규표현식으로 JSON 객체 추출 (설명 텍스트 제거)
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            json_str = match.group(0)
            data = json.loads(json_str)
        else:
            # JSON 패턴을 못 찾은 경우
            raise ValueError("No JSON object found in response")
            
    except (json.JSONDecodeError, ValueError) as e:
        st.error(f"Gemini 응답 오류: JSON 파싱 실패. 오류: {e}\n응답 내용: {text[:500]}...")
        raise
    
    # DataFrame 변환
    raw_columns = data.get("columns", [])
    raw_rows = data.get("rows", [])
    
    # 1. 컬럼 이름 생성 (중복 허용, 나중에 병합됨)
    column_names = []
    for col in raw_columns:
        sub = col.get("sub_agency", "공통")
        cond = col.get("condition", "조건")
        plan = col.get("plan", "표준")
        
        # [Hardcoded Fix] T우주 -> 5GX 프리미엄(T우주)
        if "T우주" in plan:
            plan = "5GX 프리미엄(T우주)"
            
        column_names.append(f"{sub}|{cond}({plan})")
        
    # 2. 행 데이터 -> 딕셔너리 리스트 변환 (중복 컬럼 병합)
    data_dicts = []
    for r in raw_rows:
        if not r: continue
        
        # 행 데이터 Sanitization
        sanitized_r = []
        for cell in r:
            if isinstance(cell, (dict, list)):
                sanitized_r.append(str(cell))
            else:
                sanitized_r.append(cell)
        
        # 첫 번째 값은 모델명
        model_name = str(sanitized_r[0]) if len(sanitized_r) > 0 and sanitized_r[0] is not None else "Unknown"
        row_dict = {"Model": model_name}
        
        # 나머지 값들은 가격
        values = sanitized_r[1:]
        for i, val in enumerate(values):
            if i < len(column_names):
                col_name = column_names[i]
                # 값이 유효한 경우에만 저장 (None, 빈 문자열 제외)
                if val is not None and val != "":
                    # 이미 값이 있으면? (중복 컬럼) -> 덮어쓰기
                    # (보통 Sparse해서 겹치지 않거나, 뒤에 나오는 값이 최신/유효값일 확률 높음)
                    row_dict[col_name] = val
                    
        data_dicts.append(row_dict)
        
    # 3. DataFrame 생성
    if data_dicts:
        df = pd.DataFrame(data_dicts)
        # Model 컬럼이 맨 앞에 오도록 보장 (딕셔너리 순서가 보장되지만 명시적으로)
        cols = ["Model"] + [c for c in df.columns if c != "Model"]
        df = df[cols]
    else:
        df = pd.DataFrame(columns=["Model", "Price"])
        
    # Footer Sanitization
    footer = data.get("footer", "")
    if isinstance(footer, (dict, list)):
        footer = str(footer)
    
    # 모델 코드를 표준 모델명으로 매핑
    def map_model_code_to_name(code):
        """모델 코드(SM-XXXX)를 reference_db.json의 표준 모델명으로 변환"""
        if not code or not isinstance(code, str):
            return code
        
        # 정확한 매칭 시도
        for model_info in REFERENCE_DATA.get('models', []):
            if code in model_info.get('codes', []):
                return model_info['name']
        
        # 매칭 실패시 원래 값 반환
        return code
    
    # 첫 번째 컬럼(모델명)을 표준 이름으로 변환
    if not df.empty:
        first_col = df.columns[0]
        # 첫 번째 컬럼의 값들도 문자열로 변환 (안전장치)
        df[first_col] = df[first_col].astype(str).apply(map_model_code_to_name)
        
        # 인덱스 설정 (첫 열 기준)
        # 주의: 중복된 모델명이 있을 수 있음 (다른 섹션). 따라서 인덱스로 설정하되 중복 허용
        df.set_index(first_col, inplace=True)
        
        # 전체 숫자 변환 시도
        df = df.apply(pd.to_numeric, errors='coerce')
    
    # 분석 결과만 반환 (PolicyData 객체 생성은 호출 측에서)
    return df, footer

# --- 2. 엑셀 생성 (전쟁 로직) ---
# --- 2. 엑셀 생성 (전쟁 로직) ---
def create_battle_excel(policies):
    wb = Workbook()
    
    # 1. 시트 생성
    ws_main = wb.active
    ws_main.title = "🏆최고의 정책서"
    
    # --- [New] 대리점별 추가정책 입력칸 생성 (Row 1~2) ---
    # Row 1: 대리점명
    # Row 2: 추가정책 값 (기본 0)
    # Map: policy_name -> cell_coordinate (e.g., "AgencyA" -> "$B$2")
    
    agency_adj_map = {}
    current_adj_col = 2
    
    ws_main.cell(row=1, column=1, value="대리점 추가정책")
    ws_main.cell(row=2, column=1, value="입력값(원)")
    
    for p in policies:
        cell_name = ws_main.cell(row=1, column=current_adj_col, value=p.name)
        cell_val = ws_main.cell(row=2, column=current_adj_col, value=0) # 기본값 0
        
        # 스타일링
        cell_name.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 노란색
        cell_name.alignment = Alignment(horizontal='center')
        cell_val.alignment = Alignment(horizontal='center')
        
        # 좌표 저장 (절대참조)
        col_letter = cell_val.column_letter
        agency_adj_map[p.name] = f"${col_letter}$2"
        
        current_adj_col += 1
        
    # 메인 테이블 시작 Row
    start_row = 4
    
    # --- 동적 통합 로직 시작 ---
    all_models = set()
    
    for p in policies:
        if p.df is not None and not p.df.empty:
            # 사용자가 선택한 모델만 수집 (없으면 전체)
            models_to_scan = p.selected_models if p.selected_models else p.df.index
            
            # 인덱스(모델명) 수집: 문자열로 변환하여 추가
            for idx in models_to_scan:
                if isinstance(idx, (str, int, float)):
                    val_str = str(idx).strip()
                    if val_str and val_str.lower() not in ["unknown", "none", "nan"]:
                        all_models.add(val_str)
                else:
                    all_models.add(str(idx))
            
    sorted_models = sorted([m for m in all_models if m], key=str)
    combined_index = sorted_models
    # --- 동적 통합 로직 끝 ---
    
    # --- 헤더 작성 (4대 핵심 정책 + 요금제) ---
    # 순서: 모델명, 공시(MNP), 선약(MNP), 공시(기변), 선약(기변)
    headers = [
        "모델명", 
        "공시(MNP)", "공시(MNP)요금제", 
        "선약(MNP)", "선약(MNP)요금제", 
        "공시(기변)", "공시(기변)요금제", 
        "선약(기변)", "선약(기변)요금제"
    ]
    
    for c_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=start_row, column=c_idx, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.font = Font(bold=True)
        
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Row 순회 (모델별)
    for r_idx, model in enumerate(combined_index, start_row + 1):
        ws_main.cell(row=r_idx, column=1, value=model).border = thin_border
        
        # 4대 카테고리별 최대값 및 요금제 초기화
        # 구조: {category: (max_price, best_plan, color_hex, policy_name)}
        best_values = {
            "공시(MNP)": (-1, "", None, None),
            "선약(MNP)": (-1, "", None, None),
            "공시(기변)": (-1, "", None, None),
            "선약(기변)": (-1, "", None, None)
        }
        
        # 모든 정책서 스캔
        for p in policies:
            if p.df is not None and model in p.df.index:
                # 사용자가 선택한 모델에 포함되어 있는지 확인
                if p.selected_models and model not in p.selected_models:
                    continue
                    
                # 사용자가 선택한 컬럼만 스캔
                cols_to_scan = p.selected_columns if p.selected_columns else p.df.columns
                
                # 해당 모델의 선택된 컬럼(조건) 확인
                for col in cols_to_scan:
                    col_str = str(col)
                    val = p.df.loc[model, col]
                    
                    # 값이 숫자인지 확인
                    try:
                        price = float(val)
                    except (ValueError, TypeError):
                        continue
                        
                    # 카테고리 및 요금제 파싱
                    # col_str format: "Sub|Cond(Plan)"
                    category = None
                    plan_name = ""
                    
                    # 요금제 추출 (괄호 안의 내용)
                    if "(" in col_str and ")" in col_str:
                        try:
                            plan_name = col_str.split("(")[-1].replace(")", "")
                        except:
                            plan_name = "Unknown"
                    
                    if "공시" in col_str:
                        if "MNP" in col_str:
                            category = "공시(MNP)"
                        elif "기변" in col_str:
                            category = "공시(기변)"
                    elif "선약" in col_str:
                        if "MNP" in col_str:
                            category = "선약(MNP)"
                        elif "기변" in col_str:
                            category = "선약(기변)"
                    
                    # 분류된 카테고리가 있으면 최대값 비교 및 갱신
                    if category:
                        current_max, _, _, _ = best_values[category]
                        if price > current_max:
                            best_values[category] = (price, plan_name, p.color_hex, p.name)
                            
        # 결과 작성
        # categories 순서와 headers 순서 매핑 필요
        target_categories = ["공시(MNP)", "선약(MNP)", "공시(기변)", "선약(기변)"]
        
        current_col = 2
        for cat in target_categories:
            price, plan, color, p_name = best_values[cat]
            
            # 가격 셀
            cell_price = ws_main.cell(row=r_idx, column=current_col)
            cell_price.border = thin_border
            cell_price.alignment = center_align
            
            # 요금제 셀
            cell_plan = ws_main.cell(row=r_idx, column=current_col + 1)
            cell_plan.border = thin_border
            cell_plan.alignment = center_align
            
            if price != -1:
                # [New] 수식 적용: =기본값 + 대리점추가정책셀
                if p_name and p_name in agency_adj_map:
                    adj_cell_ref = agency_adj_map[p_name]
                    cell_price.value = f"={price}+{adj_cell_ref}"
                else:
                    cell_price.value = price
                
                cell_plan.value = plan
                
                # 배경색 적용 (가격 셀에만)
                if color:
                    # #RRGGBB -> RRGGBB
                    clean_hex = color.lstrip('#')
                    if len(clean_hex) == 6:
                        cell_price.fill = PatternFill(start_color=clean_hex, end_color=clean_hex, fill_type="solid")
            else:
                cell_price.value = "" 
                cell_plan.value = ""
            
            current_col += 2

    # 4. 하단 조건문 동적 조립
    current_row = len(combined_index) + start_row + 2
    ws_main.cell(row=current_row, column=1, value="[가입 조건 및 유의사항]")
    current_row += 1
    
    for p in policies:
        if p.footer_text:
            ws_main.cell(row=current_row, column=1, value=f"■ {p.name}: {p.footer_text}")
            current_row += 1
            
    # 5. 원본 데이터 시트 (수식 적용)
    for p in policies:
        ws_raw = wb.create_sheet(title=f"원본_{p.name}")
        
        # [New] 전체 추가정책 입력칸
        ws_raw.cell(row=1, column=1, value="전체 추가정책")
        ws_raw.cell(row=1, column=2, value="입력값(원)")
        ws_raw.cell(row=1, column=3, value=0) # C1: 입력값
        adj_cell_ref = "$C$1"
        
        # 스타일링
        ws_raw.cell(row=1, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 데이터프레임 헤더 쓰기
        rows = list(dataframe_to_rows(p.df, index=True, header=True))
        # rows[0] is empty (index header placeholder)
        # rows[1] is header
        
        start_row_raw = 3
        
        # 헤더 쓰기 (Row 3)
        for c_idx, val in enumerate(rows[1], 1):
            ws_raw.cell(row=start_row_raw, column=c_idx, value=val)
            
        # 데이터 쓰기 (Row 4~)
        for r_idx, row_data in enumerate(rows[2:], start_row_raw + 1):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_raw.cell(row=r_idx, column=c_idx)
                
                # 첫 번째 컬럼(모델명)은 그대로
                if c_idx == 1:
                    cell.value = val
                else:
                    # 가격 컬럼은 수식 적용
                    try:
                        if val is not None and val != "":
                            float_val = float(val)
                            cell.value = f"={float_val}+{adj_cell_ref}"
                        else:
                            cell.value = val
                    except:
                        cell.value = val

        last_row = start_row_raw + len(rows) - 2
        ws_raw.cell(row=last_row + 2, column=1, value="조건문 원본:")
        ws_raw.cell(row=last_row + 3, column=1, value=p.footer_text)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 1. 설정 및 비밀키 관리 ---
st.set_page_config(page_title="성지당 시세표 변환기", layout="wide")

# (실제 배포시에는 st.secrets를 사용하세요. 로컬 테스트용으로 사이드바 입력)
with st.sidebar:
    st.header("🔐 서버 설정")
    
    # 기본값 설정 (Secrets에서 가져오기)
    gemini_api_key = st.secrets.get("GEMINI_API_KEY", "")
    supabase_url = st.secrets.get("SUPABASE_URL", "")
    supabase_key = st.secrets.get("SUPABASE_KEY", "")

    # Secrets가 있으면 입력창 숨김, 없으면 입력창 표시
    if gemini_api_key and supabase_url and supabase_key:
        st.success("✅ 서버 설정이 완료되었습니다.")
    else:
        if not gemini_api_key:
            gemini_api_key = st.text_input("Gemini API Key", type="password")
        if not supabase_url:
            supabase_url = st.text_input("Supabase Project URL")
        if not supabase_key:
            supabase_key = st.text_input("Supabase Anon Key", type="password")
    
    st.divider()
    
    # 모델 목록 가져오기 및 드롭다운 구성
    # 모델 목록 가져오기 및 드롭다운 구성
    # gemini-2.5-flash를 무조건 기본값(첫번째)으로 설정
    base_models = ["gemini-2.5-flash", "gemini-1.5-flash", "gemini-1.5-pro", "gemini-pro-vision"]
    model_options = ["gemini-2.5-flash"] # 시작은 flash로
    
    try:
        if gemini_api_key:
            genai.configure(api_key=gemini_api_key)
            # API에서 실제 사용 가능한 모델 리스트 가져오기
            fetched_models = [m.name.replace("models/", "") for m in genai.list_models() if 'generateContent' in m.supported_generation_methods]
            
            # fetched_models에 있는 것들을 추가하되, 중복 제거
            for m in fetched_models:
                if m not in model_options:
                    model_options.append(m)
            
            # 만약 API 호출 실패했거나 목록이 비었으면 기본 목록 사용
            if len(model_options) == 1: # flash만 있는 경우
                 for m in base_models:
                     if m not in model_options:
                         model_options.append(m)
                         
    except Exception:
        # API 키 오류시 기본 목록 사용
        model_options = base_models

    # gemini-1.5-flash가 항상 0번 인덱스에 있으므로 index=0
    model_name = st.selectbox("Gemini 모델 선택", model_options, index=0)

    st.divider()
    margin_default = st.number_input("기본 마진 설정 (단위:만원)", value=0)

# --- 2. 엑셀 생성 함수 (사용자 요청 스타일 적용) ---
def create_excel_bytes(data_json, margin_val):
    wb = Workbook()
    ws = wb.active
    ws.title = "성지 통합 시세표"

    # 스타일 정의
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 1. 상단 시세표 그리기
    top_headers = ["모델","출고가","공시지원금","SK_번이","SK_기변","SK_카드_번이","SK_카드_기변","KT_번이","KT_기변","KT_카드_번이","KT_카드_기변","LG_번이","LG_기변","LG_카드_번이","LG_카드_기변"]
    
    for col_idx, text in enumerate(top_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    current_row = 2
    top_data = data_json.get("top_data", [])
    
    if top_data:
        for row_data in top_data:
            # 데이터 길이가 헤더보다 짧을 경우를 대비해 패딩
            row_data = row_data + [None] * (len(top_headers) - len(row_data))
            
            # 앞 3열 (모델, 출고가, 공시지원금) - 그대로 출력
            for c in range(3):
                cell = ws.cell(row=current_row, column=c+1, value=row_data[c])
                cell.alignment = center_align
                cell.border = thin_border
            
            # 나머지 열 (가격 정보) - 마진 수식 적용
            for c in range(3, 15):
                val = row_data[c]
                cell = ws.cell(row=current_row, column=c+1)
                
                # 숫자인 경우에만 수식 적용, 아니면 그대로 값 출력
                if isinstance(val, (int, float)):
                    cell.value = f"={val}-$Q$2"
                elif val is not None and str(val).replace('-','').isdigit(): # 문자열이지만 숫자인 경우
                     cell.value = f"={val}-$Q$2"
                else:
                    cell.value = val if val is not None else ""
                    
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

    # 2. 중간 안내 문구
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
    msg_cell = ws.cell(row=current_row, column=1, value="위 표시 금액은 현금완납가격 입니다. 카드결제도 가능합니다.")
    msg_cell.font = Font(color="FF0000", bold=True, size=14)
    msg_cell.alignment = center_align
    current_row += 2

    # 3. 하단 조건표 그리기
    bottom_headers = ["통신사", "부가서비스조건", "월요금", "유지기간", "미가입시추가금"]
    bottom_col_ranges = [(1,3), (4,8), (9,10), (11,12), (13,15)] # 열 병합 범위
    
    # 헤더 출력
    for idx, (sc, ec) in enumerate(bottom_col_ranges):
        ws.merge_cells(start_row=current_row, start_column=sc, end_row=current_row, end_column=ec)
        cell = ws.cell(row=current_row, column=sc, value=bottom_headers[idx])
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.font = header_font
        cell.alignment = center_align
        for c in range(sc, ec+1):
            ws.cell(row=current_row, column=c).border = thin_border
    current_row += 1

    # 데이터 출력
    bottom_data = data_json.get("bottom_data", [])
    start_data_row = current_row
    
    if bottom_data:
        for row_data in bottom_data:
            # 데이터 패딩
            row_data = row_data + [""] * (len(bottom_headers) - len(row_data))
            
            for idx, (sc, ec) in enumerate(bottom_col_ranges):
                ws.merge_cells(start_row=current_row, start_column=sc, end_row=current_row, end_column=ec)
                cell = ws.cell(row=current_row, column=sc, value=row_data[idx])
                cell.alignment = center_align
                for c in range(sc, ec+1):
                    ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1
        
        # 통신사별 병합 (데이터가 10줄이라고 가정하고 3/3/4 등으로 나눔, 혹은 데이터 내용 기반)
        # 여기서는 사용자가 준 예시처럼 SK(3줄), KT(3줄), LG(3줄) 정도로 가정하되, 
        # 실제 데이터가 가변적일 수 있으므로 통신사 텍스트가 같은 것끼리 묶는 로직이 이상적이나
        # 우선 사용자 예시 코드의 하드코딩된 병합 로직을 최대한 따르되 안전장치 추가
        
        # (간단히 3등분 로직 대신, 첫번째 컬럼 값이 같으면 병합하는 로직은 복잡하므로 
        #  사용자 예시처럼 SK/KT/LG 순서대로 데이터가 온다고 가정하고 렌더링)
        pass 

    # 4. 맨 밑 유의사항 추가
    current_row += 1 
    footer_font = Font(size=9, color="333333") 
    footer_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 
    footer_align = Alignment(horizontal='center', vertical='center', wrap_text=True) 

    footer_lines = data_json.get("footer_lines", [])
    if footer_lines:
        for line in footer_lines:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
            cell = ws.cell(row=current_row, column=1, value=line)
            cell.font = footer_font
            cell.fill = footer_fill
            cell.alignment = footer_align
            
            for c in range(1, 16):
                ws.cell(row=current_row, column=c).border = thin_border
            current_row += 1

    # 5. 마진 설정 컨트롤러
    ws['Q1'] = "추가 마진 설정(만원)"
    ws['Q1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws['Q1'].font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions['Q'].width = 20
    ws['Q2'] = margin_val
    ws['Q2'].alignment = center_align
    ws['Q2'].font = Font(bold=True, size=14)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# --- 3. 메인 UI ---
st.title("📱 성지당 시세표 AI 변환 시스템")
st.caption("Powered by Gemini 3.0 & Supabase")

# 탭 구성
tab1, tab2 = st.tabs(["시세표 to 엑셀", "최고의 정책서 만들기"])

# --- Tab 1: 시세표 to 엑셀 (기존 기능) ---
with tab1:
    st.header("📸 이미지로 엑셀 만들기")
    uploaded_file = st.file_uploader("시세표 이미지를 올려주세요", type=['png', 'jpg', 'jpeg'])

    if uploaded_file and gemini_api_key and supabase_url and supabase_key:
        
        # Supabase 클라이언트 연결
        try:
            supabase: Client = create_client(supabase_url, supabase_key)
        except Exception as e:
            st.error(f"Supabase 연결 오류: {e}")
            st.stop()
        
        if st.button("AI 변환 시작"):
            with st.status("작업을 진행하고 있습니다...", expanded=True) as status:
                
                # 1. Supabase Storage에 원본 이미지 업로드 (uploads 버킷)
                status.write("1️⃣ 원본 이미지를 서버에 저장 중...")
                file_bytes = uploaded_file.getvalue()
                # 한글 파일명 등으로 인한 오류 방지를 위해 UUID 사용
                file_ext = uploaded_file.name.split('.')[-1]
                file_name = f"simple-ocr/{int(time.time())}_{uuid.uuid4()}.{file_ext}"
                
                try:
                    # Storage 버킷 이름: uploads
                    supabase.storage.from_("uploads").upload(file_name, file_bytes, {"content-type": uploaded_file.type})
                    # 공개 URL 가져오기
                    image_public_url = supabase.storage.from_("uploads").get_public_url(file_name)
                except Exception as e:
                    error_msg = str(e)
                    if "Bucket not found" in error_msg or "404" in error_msg:
                        st.error("❌ **오류: 'uploads' 버킷을 찾을 수 없습니다.**")
                        st.info("Supabase 대시보드 > Storage 메뉴로 이동해서 **'uploads'** 라는 이름의 **Public Bucket**을 새로 만들어주세요.")
                    elif "row-level security policy" in error_msg or "403" in error_msg:
                        st.error("❌ **오류: 권한이 없습니다 (RLS Policy).**")
                        st.info("Supabase Storage의 'uploads' 버킷에 대해 Public Access 정책을 설정해주세요.")
                    else:
                        st.error(f"이미지 업로드 실패: {e}")
                    st.stop()
                    
                # 2. Gemini 3.0 호출 (OCR)
                status.write(f"2️⃣ Gemini ({model_name})가 데이터를 추출 중...")
                genai.configure(api_key=gemini_api_key)
                # 사용자가 선택한 모델 사용
                model = genai.GenerativeModel(model_name) 
                
                prompt = """
                Analyze the provided price sheet image and extract data into a specific JSON structure.
                
                The JSON must have these keys: "top_data", "bottom_data", "footer_lines".

                1. "top_data": A list of lists representing the main price table.
                   - Columns should correspond to: [Model, FactoryPrice, PublicSupport, SK_Move, SK_Change, SK_Card_Move, SK_Card_Change, KT_Move, KT_Change, KT_Card_Move, KT_Card_Change, LG_Move, LG_Change, LG_Card_Move, LG_Card_Change]
                   - Extract numerical values for prices. If a cell is empty or has '-', use null or 0.
                   - Example row: ["Flip7 256", 148.5, 60, 13, 18, -27, -22, 15, 15, -25, -25, -3, -1, -43, -41]

                2. "bottom_data": A list of lists for the carrier condition table at the bottom.
                   - Columns: [Carrier, ServiceCondition, MonthlyFee, Duration, Penalty]
                   - Example row: ["SK(24months)", "Plan: Premium", "109,000won", "6 months", "500,000"]

                3. "footer_lines": A list of strings for the caution/notice text at the very bottom.
                   - Capture each distinct line of text as a string in the list.
                
                Output ONLY valid JSON.
                """
                
                # 재시도 로직 추가 (429 Rate Limit 대응)
                max_retries = 3
                retry_delay = 5 # 초
                
                for attempt in range(max_retries):
                    try:
                        response = model.generate_content([prompt, {"mime_type": uploaded_file.type, "data": file_bytes}])
                        
                        # JSON 파싱
                        json_str = response.text.replace("```json", "").replace("```", "").strip()
                        data_json = json.loads(json_str)
                        break # 성공하면 루프 탈출
                    except Exception as e:
                        error_msg = str(e)
                        if "429" in error_msg and attempt < max_retries - 1:
                            status.write(f"⚠️ 사용량 초과(429). {retry_delay}초 후 재시도합니다... ({attempt+1}/{max_retries})")
                            time.sleep(retry_delay)
                            retry_delay *= 2 # 대기 시간 2배로 늘림
                        else:
                            st.error(f"Gemini 처리 실패: {e}")
                            st.stop()
                
                # 3. 엑셀 파일 생성
                status.write("3️⃣ 엑셀 파일 생성 중...")
                excel_bytes = create_excel_bytes(data_json, margin_default)
                
                # 4. 엑셀 파일 Supabase 저장 (exports 버킷)
                status.write("4️⃣ 엑셀 파일을 클라우드에 백업 중...")
                excel_name = f"simple-excel/converted_{int(time.time())}.xlsx"
                try:
                    supabase.storage.from_("exports").upload(excel_name, excel_bytes.getvalue(), {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
                    excel_public_url = supabase.storage.from_("exports").get_public_url(excel_name)
                except Exception as e:
                    error_msg = str(e)
                    if "Bucket not found" in error_msg or "404" in error_msg:
                        st.error("❌ **오류: 'exports' 버킷을 찾을 수 없습니다.** (엑셀 저장 실패)")
                        st.info("Supabase 대시보드에서 'exports' 버킷을 생성했는지 확인해주세요.")
                    else:
                        st.error(f"엑셀 업로드 실패: {e}")
                    st.stop()

                # 5. DB에 기록 남기기
                status.write("5️⃣ 작업 이력 기록 중...")
                try:
                    supabase.table("price_sheets").insert({
                        "filename": uploaded_file.name,
                        "image_url": image_public_url,
                        "excel_url": excel_public_url,
                        "status": "success"
                    }).execute()
                except Exception as e:
                    st.warning(f"DB 기록 실패 (파일은 생성됨): {e}")
                
                status.update(label="완료되었습니다!", state="complete", expanded=False)

            # 결과 화면
            st.success("변환 성공!")
            col1, col2 = st.columns(2)
            with col1:
                st.image(uploaded_file, caption="원본 이미지")
            with col2:
                st.info("생성된 엑셀 파일")
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=excel_bytes,
                    file_name=excel_name.split('/')[-1],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.markdown(f"[클라우드 링크로 보기]({excel_public_url})")

    elif not (gemini_api_key and supabase_url):
        st.warning("왼쪽 사이드바에서 서버 설정(API Key)을 완료해주세요.")

# --- Tab 2: 최고의 정책서 만들기 (커스텀 정책 배틀) ---
with tab2:
    st.header("⚔️ 성지당 v2: 커스텀 정책 배틀")
    st.markdown("대리점 이름과 색상을 직접 정해서 **최고의 정책서**를 만들어보세요.")
    st.caption("데이터는 'uploads' 및 'exports' 버킷에 체계적으로 분류되어 저장됩니다.")

    if 'policies' not in st.session_state:
        st.session_state.policies = []
    
    # 색상 상태 관리 (파일 업로드 시에는 변경되지 않음)
    if 'current_color' not in st.session_state:
        st.session_state.current_color = get_random_pastel_color()

    # 탭 2 내부에 별도의 입력 구역 생성 (사이드바 대신)
    with st.expander("➕ 새로운 경쟁자 등록하기", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            input_agency_name = st.text_input("대리점 이름 (예: 구로 1호점)", placeholder="이름을 지어주세요")
            # 현재 세션에 저장된 색상 사용
            input_agency_color = st.color_picker("고유 색상 선택", st.session_state.current_color)
        with col2:
            uploaded_battle_file = st.file_uploader("시세표 이미지 업로드 (배틀용)", type=['png', 'jpg'], key="battle_uploader")
        
        if st.button("목록에 추가 +", type="primary"):
            if uploaded_battle_file and input_agency_name:
                file_bytes = uploaded_battle_file.getvalue()
                
                # AI 분석 없이 이미지와 메타데이터만 저장
                policy_data = PolicyData(
                    name=input_agency_name,
                    image_bytes=file_bytes,
                    color_hex=input_agency_color
                )
                
                st.session_state.policies.append(policy_data)
                st.success(f"✅ '{input_agency_name}' 목록에 추가 완료! (분석은 Battle Start 시 진행됩니다)")
                
                # 성공적으로 추가된 후에만 색상 변경
                st.session_state.current_color = get_random_pastel_color()
                
                
            elif not input_agency_name:
                st.error("대리점 이름을 입력해주세요!")
            elif not uploaded_battle_file:
                st.error("시세표 이미지를 업로드해주세요!")

    # 메인 화면: 현황판
    st.subheader(f"🥊 참전 대기 중인 대리점: {len(st.session_state.policies)}곳")

    if len(st.session_state.policies) > 0:
        cols = st.columns(4)
        for idx, p in enumerate(st.session_state.policies):
            with cols[idx % 4]:
                status_icon = "⏳" if not p.is_analyzed else "✅"
                st.markdown(
                    f"""
                    <div style='background-color: {p.color_hex}; padding: 15px; border-radius: 10px; margin-bottom: 10px;'>
                        <h4 style='margin: 0; color: #333;'>{status_icon} {p.name}</h4>
                        <p style='margin: 5px 0 0 0; font-size: 0.9em; color: #555;'>대기 중...</p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                # 삭제 버튼
                if st.button(f"🗑️ 삭제", key=f"delete_{idx}"):
                    st.session_state.policies.pop(idx)

    # 메인 화면: 현황판
    st.subheader(f"🥊 현재 참전 중인 대리점: {len(st.session_state.policies)}곳")

    if len(st.session_state.policies) > 0:
        cols = st.columns(4)
        for idx, p in enumerate(st.session_state.policies):
            with cols[idx % 4]:
                status_icon = "⏳" if not p.is_analyzed else "✅"
                model_count = f"모델 {len(p.df)}개" if p.is_analyzed else "대기 중..."
                # 카드를 해당 색상으로 꾸미기
                st.markdown(
                    f"""
                    <div style="
                        background-color: {p.color_hex};
                        padding: 15px;
                        border-radius: 10px;
                        border: 1px solid #ddd;
                        color: black;
                        text-align: center;
                        box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
                    ">
                        <h4 style="margin:0; color:black;">{status_icon} {p.name}</h4>
                        <p style="margin:0; font-size:0.8em;">{model_count}</p>
                    </div>
                    """, 
                    unsafe_allow_html=True
                )
                # 조건문 미리보기 (분석 완료된 경우만)
                if p.is_analyzed and p.footer_text:
                    with st.expander("조건 보기"):
                        st.text(p.footer_text[:100] + "...")

        st.divider()

        # 엑셀 생성 버튼 영역
        col1, col2 = st.columns([1, 2])
        with col1:
            # 1단계: AI 분석 시작
            if st.button("🚀 1. AI 분석 시작 (Analysis Start)", type="primary", use_container_width=True):
                with st.spinner("🤖 AI가 모든 시세표를 분석 중..."):
                    for idx, policy in enumerate(st.session_state.policies):
                        if not policy.is_analyzed:
                            try:
                                # Gemini 분석
                                df, footer_text = parse_image_with_gemini_v2(
                                    policy.image_bytes, 
                                    policy.name, 
                                    policy.color_hex, 
                                    gemini_api_key, 
                                    model_name
                                )
                                # 결과를 현재 policy 객체에 업데이트
                                policy.df = df
                                policy.footer_text = footer_text
                                policy.is_analyzed = True
                                
                                # 초기 선택값 설정 (전체 선택)
                                if df is not None:
                                    policy.selected_models = df.index.tolist()
                                    policy.selected_columns = df.columns.tolist()
                                
                                # Supabase에 이미지 업로드 및 DB 저장
                                if supabase_url and supabase_key:
                                    try:
                                        supabase_v2: Client = create_client(supabase_url, supabase_key)
                                        file_name = f"policy-battle/{int(time.time())}_{uuid.uuid4()}.jpg"
                                        
                                        supabase_v2.storage.from_("uploads").upload(
                                            file_name, 
                                            policy.image_bytes, 
                                            {"content-type": "image/jpeg"}
                                        )
                                        image_url = supabase_v2.storage.from_("uploads").get_public_url(file_name)
                                        
                                        # DB에 로그 저장
                                        parsed_json = policy.df.to_json(orient='split', force_ascii=False)
                                        supabase_v2.table("policy_uploads").insert({
                                            "agency_name": policy.name,
                                            "image_url": image_url,
                                            "parsed_data": json.loads(parsed_json)
                                        }).execute()
                                    except Exception as e:
                                        st.warning(f"'{policy.name}' 클라우드 저장 실패: {e}")
                                
                                st.toast(f"✅ {policy.name} 분석 완료!", icon="✅")
                                
                            except Exception as e:
                                st.error(f"'{policy.name}' 분석 실패: {e}\n\n{traceback.format_exc()}")
                                # 실패해도 계속 진행
                    
                    st.success("AI 분석이 완료되었습니다! 아래에서 데이터를 검토해주세요.")
                    st.session_state['analysis_done'] = True

        # 2단계: 검토 및 엑셀 생성 (분석 완료 시 표시)
        analyzed_policies = [p for p in st.session_state.policies if p.is_analyzed]
        
        if analyzed_policies:
            st.divider()
            st.subheader("🧐 데이터 검토 및 필터링")
            st.info("각 대리점 탭을 눌러서 제외하고 싶은 모델(행)이나 조건(열)을 체크 해제하세요.")
            
            # 대리점별 탭 생성
            tabs = st.tabs([p.name for p in analyzed_policies])
            
            for idx, p in enumerate(analyzed_policies):
                # 하위 호환성: id가 없는 기존 객체에 id 부여
                if not hasattr(p, 'id'):
                    p.id = str(uuid.uuid4())
                    
                with tabs[idx]:
                    if p.df is not None and not p.df.empty:
                        c1, c2 = st.columns([1, 3])
                        with c1:
                            st.markdown(f"**[{p.name}] 필터 설정**")
                            # 모델(행) 선택
                            selected_rows = st.multiselect(
                                f"포함할 모델 ({len(p.df)}개)",
                                options=p.df.index.tolist(),
                                default=p.selected_models if p.selected_models else p.df.index.tolist(),
                                key=f"rows_{p.id}"
                            )
                            # 조건(열) 선택
                            selected_cols = st.multiselect(
                                f"포함할 조건 ({len(p.df.columns)}개)",
                                options=p.df.columns.tolist(),
                                default=p.selected_columns if p.selected_columns else p.df.columns.tolist(),
                                key=f"cols_{p.id}"
                            )
                            
                            # 선택 상태 업데이트
                            p.selected_models = selected_rows
                            p.selected_columns = selected_cols
                            
                        with c2:
                            st.markdown("**데이터 미리보기** (선택된 항목만 엑셀에 반영됩니다)")
                            # 필터링된 데이터프레임 보여주기
                            try:
                                filtered_df = p.df.loc[selected_rows, selected_cols]
                                st.dataframe(filtered_df, use_container_width=True)
                            except Exception as e:
                                st.error(f"데이터 표시 오류: {e}")
                    else:
                        st.warning("분석된 데이터가 없습니다.")

            st.divider()
            
            # 3단계: 최종 엑셀 생성 버튼
            if st.button("📊 2. 최고의 정책서 만들기 (Generate Excel)", type="primary", use_container_width=True):
                with st.spinner("최종 엑셀 파일을 생성하고 있습니다..."):
                    # 엑셀 생성 (필터링된 데이터 반영은 create_battle_excel 내부에서 처리 필요)
                    excel_file = create_battle_excel(analyzed_policies)
                    st.session_state['excel_ready'] = excel_file
                    
                    # Supabase 업로드 로직 (기존과 동일)
                    if supabase_url and supabase_key:
                        try:
                            supabase_v2: Client = create_client(supabase_url, supabase_key)
                            excel_name = f"battle-results/best_policy_{int(time.time())}.xlsx"
                            
                            supabase_v2.storage.from_("exports").upload(excel_name, excel_file.getvalue(), {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
                            excel_url = supabase_v2.storage.from_("exports").get_public_url(excel_name)
                            
                            participants = [p.name for p in analyzed_policies]
                            supabase_v2.table("battle_results").insert({
                                "excel_url": excel_url,
                                "participants": participants
                            }).execute()
                            st.toast("클라우드 저장 완료!", icon="☁️")
                        except Exception as e:
                            st.warning(f"클라우드 백업 실패: {e}")
                            
                    st.success("완성되었습니다! 아래 버튼을 눌러 다운로드하세요.")

        with col2:
            if 'excel_ready' in st.session_state:
                st.download_button(
                    label="📥 결과물 다운로드 (Excel)",
                    data=st.session_state['excel_ready'],
                    file_name="성지당_최고의정책서_커스텀.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    else:
        st.info("위의 '새로운 경쟁자 등록하기'에서 대리점 이름과 이미지를 넣고 '추가' 버튼을 눌러주세요.")
